import os
import glob
from docx import Document
from openpyxl import load_workbook
import re
from PyPDF2 import PdfReader, PdfFileReader
from concurrent.futures import ThreadPoolExecutor

rootDir = r"C:/Users/Admin/Documents"
search_term = os.path.join(rootDir ,"historia *")
files = glob.glob(search_term)

print("Znalezione:")
for file in files:
    print(file)

phrase = input("Podaj regex do wyszukania: ").lower()

try:
    pattern = re.compile(phrase, re.IGNORECASE)

except re.error as e:
    print(f"Niepoprawny regex: {phrase}")
    exit(1)

def search_xlsx(path,regex):
    try:
        wb = load_workbook(path,data_only=True)

        for sheet in wb.worksheets:
            for r,row in enumerate(sheet.iter_rows(values_only=True), start=1):
                for c, cell in enumerate(row, start=1):

                    if cell is None:
                        continue
                    text =str(cell)

                    if regex.search(text):
                        print(
                        f"📊 {path} | arkusz '{sheet.title}' "
                        f"| komórka {r}:{c}: {text}"
                        )
    except Exception as e:
        print(f"Błąd XLSX {path}: error: {e}")

def search_pdf(path):
    reader = PdfFileReader(path)

    for i,page in enumerate(reader.pages, start=1):
        text=page.extractText() or ""

        if  regex.search(text):
            print(f" PDF {path} | strona {i}")

def process_file(path):
    ext = os.path.splitext(path)[1].lower()

    try:
        if ext == ".xlsx":
            search_xlsx(path)
        elif ext == ".pdf":
            search_pdf(path)
    except Exception as e:
        print(f"⚠️ Błąd {path}: {e}")

files_to_process = []

for root, dirs, files in os.walk(rootDir):

    for name in files:
        # ignorowanie wielkości liter
        if name.lower().startswith(search_term.lower()):
            files_to_process.append(os.path.join(root, name))

with ThreadPoolExecutor() as executor:
    executor.map(process_file, files_to_process)
# wyszukiwanie
# for root in xlsx_files:
#     wb = load_workbook(root, data_only=True)
#     for sheet in wb.worksheets:
#         for r, row in enumerate(sheet.iter_rows(values_only=True),start=1):
#             for c, cell in enumerate(row,start=1):
#                 if cell and phrase in str(cell).lower():
#                     print(sheet.cell(row=r, column=c).value)
#


# globs = glob.iglob(file_to_search,
#                    root_dir=rootDir,
#                    recursive=True)
# for i,file in enumerate(globs,start=1):
#     try:
#         with open(file) as f:
#             content = f.read()
#
#             if(search_term in content):
#                 files_to_check.append(file)
#                 print(files_to_check)
#     except:
#         pass



#dirlist = os.listdir(rootDir)

#filelist = []hd
#for item in dirlist:
    #filelist.append(os.path.join(rootDir, item))
    #print(item, "\n")
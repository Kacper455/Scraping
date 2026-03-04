import os
import glob
from docx import Document
from openpyxl import load_workbook
import pandas as pd
files_to_check = []
#key_word = input()
rootDir = r"C:/Users/Admin/Documents"
search_term = os.path.join(rootDir ,"historia *")
files = glob.glob(search_term)

print("Znalezione:")
for file in files:
    print(file)

phrase = input("Napisz czego szukasz: ").lower()

# pliki
xlsx_files = glob.glob(file)

# wyszukiwanie
for root in xlsx_files:
    wb = load_workbook(root, data_only=True)
    for sheet in wb.worksheets:
        for r, row in enumerate(sheet.iter_rows(values_only=True),start=1):
            for c, cell in enumerate(row,start=1):
                if cell and phrase in str(cell).lower():
                    print(sheet.cell(row=r, column=c).value)



# globs = glob.iglob(file_to_search,
#                    root_dir=rootDir,
#                    recursive=True)
# for i,file in enumerate(globs,start=1):
#     try:
#         with open(file) as f:
#             content = f.read()
#
#             if(search_term in content):
#                 files_to_check.append(file)
#                 print(files_to_check)
#     except:
#         pass



#dirlist = os.listdir(rootDir)

#filelist = []hd
#for item in dirlist:
    #filelist.append(os.path.join(rootDir, item))
    #print(item, "\n")
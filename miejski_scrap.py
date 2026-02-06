import requests
from urllib.parse import urljoin
from bs4 import BeautifulSoup as bs
import pandas as pd
from pathlib import Path
from openpyxl.reader.excel import load_workbook
import time

urel = 'https://www.miejski.pl/e-'

urls = []
seen = set()

for i in range(1,14):
    url = f"{urel}{i}.html"
    if url not in seen:
        seen.add(url)
        urls.append(url)

for url in urls:
    print(url)

list = []
result = []
listed = []
list2 = []

for idx, addresses in enumerate(urls, start=1):
    response = requests.get(addresses).content
    soup = bs(response, 'html.parser')

    for x in soup.find_all('footer'):
        link = x.find("a", href=True)
        l = link['href']
    length = len(l)

    for x in soup.find_all('ul', id = "simple-link-list"):
        for link in x.find_all('li'):
            ln = link.find("a", href=True)
            list.append(ln['href'])

            list = [
                urljoin(url_to_scape, a['href'])
                for a in x.find_all("a", href=True)
                if a["href"].startswith("/")]
            pr = len(list)


    for idx,address in enumerate(list, start=1):
        page = requests.get(address).content
        soup = bs(page, 'html.parser')

        for a in soup.find_all('main'):
            for art in a.find_all('article'):
                for head in art.find_all("header"):
                    header = head.find_all(string=True)

        l = listed.append(header)
        ll = len(listed)

        for x in soup.find_all('main'):
            title = x.find("div", class_="rating-box")
            rate = title.find("span", class_="rating")
            rating = rate.text
            inr = int(rating)
            result.append(inr)

        r = len(result)
    #print(header)
if (ll == r):
    data = {
    'Nazwa': listed,
    'Rating': result,
    }
    df = pd.DataFrame(data)
    df_sorted = df.sort_values(by=['Rating'], ascending=False)
    print(df_sorted)
else:
    print("Data is not matching")
    print(result)
    print(ll,r)

documents = Path.home() / "Documents"
file_path = documents / "Wyniki.xlsx"

df_sorted.to_excel(file_path, index=False, startcol=2)

if file_path.exists():
    book = load_workbook(file_path)
    start_row = book.active.max_row
else:
    start_row = 0
with pd.ExcelWriter(
    file_path,
    engine="openpyxl",
    mode="a" if file_path.exists() else "w",
    if_sheet_exists="overlay"
) as writer:
    df.to_excel(
        writer,
        index=False,
        header=False,
        startrow=start_row
    )


input("Press enter to continue...")

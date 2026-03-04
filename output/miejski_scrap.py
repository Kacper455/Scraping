import requests
from urllib.parse import urljoin
from bs4 import BeautifulSoup as bs
import pandas as pd
from pathlib import Path
from openpyxl.reader.excel import load_workbook
import psutil
import time

def getHTML(url):
    r = requests.get(url)
    r.raise_for_status()
    return r.content

urel = 'https://www.miejski.pl/h-'

urls = []
seen = set()

def urls_to_scrap():
    for i in range(1,22):
        url = f"{urel}{i}.html"
        if url not in seen:
            seen.add(url)
            urls.append(url)
    return urls

result = []
listed = []
adress = []

for idx, addresses in enumerate(urls_to_scrap(), start=1):
    response = requests.get(addresses).content
    soup = bs(response, 'html.parser')

    for x in soup.find_all('footer'):
        link = x.find("a", href=True)
        l = link['href']
    length = len(l)


    def create_links(soup, addresses):
        links = []

        for x in soup.find_all('ul', id = "simple-link-list"):
            for link in x.find_all('li'):
                ln = link.find("a", href=True)

                if ln['href'].split('/'):
                    full_url = urljoin(addresses, ln['href'])
                    links.append(full_url)
        return links

    links = create_links(soup, addresses)

    for idx,address in enumerate(links, start=1):
        page = requests.get(address).content
        soup = bs(page, 'html.parser')

        for a in soup.find_all('main'):
            for art in a.find_all('article'):
                for head in art.find_all("header"):
                    header = head.find_all(string=True)

        l = listed.append(header)
        ll = len(listed)

        for x in soup.find_all('main'):
            title = x.find("div", class_="rating-box")
            rate = title.find("span", class_="rating")
            rating = rate.text
            inr = int(rating)
            result.append(inr)

        r = len(result)

    def create_data():
        try:
            data = {
            'Nazwa': listed,
            'Rating': result,
            }
            df = pd.DataFrame(data)
            df_sorted = df.sort_values(by=['Rating'], ascending=False)

            documents = Path.home() / "Documents"
            file_path = documents / "Wyniki_H.xlsx"

            df_sorted.to_excel(file_path, index=False, startcol=2)

            if file_path.exists():
                book = load_workbook(file_path)
                start_row = book.active.max_row
            else:
                start_row = 0
            with pd.ExcelWriter(
                    file_path,
                    engine="openpyxl",
                    mode="a" if file_path.exists() else "w",
                    if_sheet_exists="overlay"
            ) as writer:
                df_sorted.to_excel(
                    writer,
                    index=False,
                    header=False,
                    startrow=start_row
                )
        except:
            print("You cannot override open file...")
            for process in psutil.process_iter():
                if process.name() == "EXCEL.EXE":
                    process.kill()

def __main__():
    create_data()

if __name__ == "__main__":
    __main__()
try:
    input("Press enter to continue...")
except:
    pass

# funkcja create_links działa poprawnie, ale nie przekazuje swojego wyniku do następnej pętli która scrapuje dane, przez co nic nie jest wysyłane do końcowego pliku .xlsx
#dodac printowanie na tablice wszystkich adresow z funkcji ulrs_to_scrap
# sprawdzic zapisywanie do pliku wyniki.xlsx
# poprawic nadpisywanie danych po kazdym wywolaniu programu dla danej litery alfabetu,albo rozszerzyc parsowanie dla kazdego kafelka kazdej litery/znaku, pod warunkiem wprowadzenia threadingu